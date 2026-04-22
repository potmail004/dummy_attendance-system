from fastapi import APIRouter
from pydantic import BaseModel
from services.db_service import get_connection
from fastapi.responses import FileResponse
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter()


class ReportRequest(BaseModel):
    class_number: str
    date: str  # yyyy-mm-dd


@router.post("/report")
def generate_report(data: ReportRequest):
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # 🔹 class details
        cursor.execute(
            """
            SELECT id, class_code, semester, slot, type
            FROM classes WHERE class_number=%s
        """,
            (data.class_number,),
        )
        class_row = cursor.fetchone()

        if not class_row:
            cursor.close()
            conn.close()
            return {"error": "Class does not exist"}

        class_id, class_code, semester, slot, type_ = class_row

        # 🔹 teacher
        cursor.execute(
            """
            SELECT t.full_name
            FROM class_teacher ct
            JOIN teachers t ON ct.teacher_id = t.id
            WHERE ct.class_id=%s
        """,
            (class_id,),
        )
        teacher_row = cursor.fetchone()
        teacher_name = teacher_row[0] if teacher_row else "N/A"

        # 🔹 students
        cursor.execute(
            """
            SELECT s.id, s.full_name, s.student_code
            FROM class_student cs
            JOIN students s ON cs.student_id = s.id
            WHERE cs.class_id=%s
            ORDER BY s.student_code
        """,
            (class_id,),
        )
        students = cursor.fetchall()

        # 🔹 session check
        cursor.execute(
            """
            SELECT id FROM attendance_sessions
            WHERE class_id=%s AND session_date=%s
        """,
            (class_id, data.date),
        )
        session_row = cursor.fetchone()

        if not session_row:
            cursor.close()
            conn.close()
            return {"error": "No attendance record found for this date"}

        session_id = session_row[0]

        # 🔹 attendance
        cursor.execute(
            """
            SELECT student_id FROM attendance_records
            WHERE session_id=%s
        """,
            (session_id,),
        )
        present_students = cursor.fetchall()

        present_set = set([p[0] for p in present_students])

        rows = []
        for s in students:
            rows.append(
                {
                    "Student Name": s[1],
                    "Student Code": s[2],
                    "Attendance": "Present" if s[0] in present_set else "Absent",
                }
            )

        df = pd.DataFrame(rows)

        # 🔹 file
        file_path = f"report_{data.class_number}_{data.date}.xlsx"

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, startrow=9)

            workbook = writer.book
            sheet = writer.sheets["Sheet1"]

            # 🔹 header
            header_data = [
                ("Teacher Name", teacher_name),
                ("Class Code", class_code),
                ("Slot", slot),
                ("Semester", semester),
                ("Class Number", data.class_number),
                ("Date", data.date),
                ("Type", type_),
            ]

            for i, (key, value) in enumerate(header_data, start=1):
                sheet.cell(row=i, column=1, value=key).font = Font(bold=True)
                sheet.cell(row=i, column=2, value=value)

            # 🔹 table header style
            header_row = 10
            for col in range(1, 4):
                cell = sheet.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 🔹 column widths
            widths = [25, 20, 15]
            for i, w in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(i)].width = w

            # 🔹 align attendance
            for row in range(header_row + 1, header_row + 1 + len(df)):
                sheet.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        cursor.close()
        conn.close()

        return FileResponse(file_path, filename=file_path)

    except Exception as e:
        return {"error": str(e)}
